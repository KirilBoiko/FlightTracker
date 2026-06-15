#!/usr/bin/env python3
"""
Ridge Regression price estimator for TBS/BUS ↔ TLV flight data.
Outputs a styled two-sheet Excel workbook.
"""

import re
import numpy as np
import pandas as pd
from sklearn.linear_model import Ridge
from sklearn.preprocessing import OneHotEncoder
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.model_selection import KFold, cross_val_predict
from sklearn.metrics import mean_absolute_error, mean_squared_error
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# ── 1. Load data ──────────────────────────────────────────────────────────
CSV = "FlightData_v2.3 - FlightData-2.csv"
df = pd.read_csv(CSV)
print(f"Loaded {len(df)} rows, {df['price_usd'].isna().sum()} missing prices.")

# ── 2. Feature engineering ────────────────────────────────────────────────
def parse_hour(t):
    """'13:30' → 13.5"""
    try:
        parts = str(t).strip().split(":")
        return int(parts[0]) + int(parts[1]) / 60
    except Exception:
        return np.nan

df["departure_hour"] = df["departure_time"].apply(parse_hour)
df["route_airline"]  = (df["origin"] + "-" + df["destination"]
                        + "_" + df["airline"])

# ── 3. Split known / unknown ──────────────────────────────────────────────
known   = df[df["price_usd"].notna()].copy()
unknown = df[df["price_usd"].isna()].copy()

FEATURES_CAT = ["route_airline"]
FEATURES_NUM = ["days_till_departure", "departure_hour", "Capacity"]
TARGET       = "price_usd"

X_known = known[FEATURES_CAT + FEATURES_NUM]
y_known = known[TARGET]
X_all   = df[FEATURES_CAT + FEATURES_NUM]

# ── 4. Build pipeline ─────────────────────────────────────────────────────
preprocessor = ColumnTransformer([
    ("ohe", OneHotEncoder(handle_unknown="ignore", sparse_output=False),
     FEATURES_CAT),
    ("num", "passthrough", FEATURES_NUM),
])

model = Pipeline([
    ("pre", preprocessor),
    ("ridge", Ridge(alpha=5)),
])

# ── 5. 5-fold CV for OOF predictions ─────────────────────────────────────
kf = KFold(n_splits=5, shuffle=True, random_state=42)
oof_preds = cross_val_predict(model, X_known, y_known, cv=kf)

cv_mae  = mean_absolute_error(y_known, oof_preds)
cv_rmse = np.sqrt(mean_squared_error(y_known, oof_preds))
print(f"CV MAE:  ${cv_mae:.2f}")
print(f"CV RMSE: ${cv_rmse:.2f}")

# Per route_airline MAE from OOF
known_oof = known.copy()
known_oof["_oof"] = oof_preds
ra_mae = (
    known_oof.groupby("route_airline")
    .apply(lambda g: mean_absolute_error(g[TARGET], g["_oof"]))
    .rename("mae")
)
print("\nPer route_airline MAE:")
print(ra_mae.round(2).to_string())

# ── 6. Train on full known set, predict all rows ──────────────────────────
model.fit(X_known, y_known)
df["_pred"] = model.predict(X_all)

# ── 7. Build output columns ───────────────────────────────────────────────
def get_error(ra):
    return round(ra_mae.get(ra, cv_mae), 2)

df["estimated_price"] = np.where(
    df["price_usd"].notna(),
    df["price_usd"],
    df["_pred"].round(2)
)
df["estimation_error"] = df.apply(
    lambda r: "actual" if pd.notna(r["price_usd"])
              else get_error(r["route_airline"]),
    axis=1
)
df["price_source"] = np.where(df["price_usd"].notna(), "actual", "estimated")

# ── 8. Write Excel ────────────────────────────────────────────────────────
OUT = "FlightData_v2.3_estimated-2.xlsx"

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
YELLOW_FILL  = PatternFill("solid", fgColor="FFFF99")
GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")
HEADER_FONT  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
CELL_FONT    = Font(name="Arial", size=10)
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left",   vertical="center")
THIN         = Side(style="thin", color="CCCCCC")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(cell, text):
    cell.value     = text
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = CENTER
    cell.border    = BORDER

def style_cell(cell, value, align=LEFT):
    cell.value     = value
    cell.font      = CELL_FONT
    cell.alignment = align
    cell.border    = BORDER

wb = openpyxl.Workbook()

# ── Sheet 1: Flights ──────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Flights"

# Columns to output (drop internal helpers)
out_cols = [c for c in df.columns
            if c not in ("_pred", "route_airline", "departure_hour", "Route")]
# Append our new columns
for nc in ["estimated_price", "estimation_error", "price_source"]:
    if nc not in out_cols:
        out_cols.append(nc)

# Header row
for ci, col in enumerate(out_cols, 1):
    style_header(ws1.cell(row=1, column=ci), col)

# Data rows
for ri, (_, row) in enumerate(df.iterrows(), 2):
    is_estimated = row["price_source"] == "estimated"
    fill = YELLOW_FILL if is_estimated else GREEN_FILL

    for ci, col in enumerate(out_cols, 1):
        val = row[col]
        # Clean up numpy types
        if isinstance(val, (np.integer,)):  val = int(val)
        elif isinstance(val, (np.floating,)): val = None if np.isnan(val) else float(val)
        elif isinstance(val, float) and np.isnan(val): val = None

        cell = ws1.cell(row=ri, column=ci)
        style_cell(cell, val)
        cell.fill = fill

# Column widths
col_widths = {
    "snapshot_date": 20, "departure_date": 14, "days_till_departure": 10,
    "origin": 8, "destination": 10, "airline": 18, "flight_number": 14,
    "departure_time": 14, "aircraft": 10, "price_usd": 12,
    "duration_minutes": 10, "is_direct": 9, "Capacity": 10,
    "estimated_price": 16, "estimation_error": 16, "price_source": 13,
}
for ci, col in enumerate(out_cols, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 13)

ws1.freeze_panes = "A2"

# ── Sheet 2: Model Summary ────────────────────────────────────────────────
ws2 = wb.create_sheet("Model Summary")

# Overall metrics table
overall_headers = ["Metric", "Value"]
for ci, h in enumerate(overall_headers, 1):
    style_header(ws2.cell(row=1, column=ci), h)

overall_rows = [
    ("Training rows (known prices)", len(known)),
    ("Rows to estimate",             len(unknown)),
    ("Total rows",                   len(df)),
    ("CV Folds",                     5),
    ("Ridge alpha",                  5),
    ("CV MAE  ($)",                  round(cv_mae, 2)),
    ("CV RMSE ($)",                  round(cv_rmse, 2)),
]
for ri, (metric, val) in enumerate(overall_rows, 2):
    style_cell(ws2.cell(row=ri, column=1), metric)
    style_cell(ws2.cell(row=ri, column=2), val, CENTER)

# Per route_airline MAE table
ws2.cell(row=10, column=1).value = ""
hdr_row = 11
ra_headers = ["Route × Airline", "MAE ($)", "Training rows", "Estimated rows"]
for ci, h in enumerate(ra_headers, 1):
    style_header(ws2.cell(row=hdr_row, column=ci), h)

ra_train_ct = known.groupby("route_airline").size()
ra_est_ct   = unknown.groupby("route_airline").size()

for ri, (ra, mae_val) in enumerate(ra_mae.sort_values().items(), hdr_row + 1):
    style_cell(ws2.cell(row=ri, column=1), ra)
    style_cell(ws2.cell(row=ri, column=2), round(mae_val, 2), CENTER)
    style_cell(ws2.cell(row=ri, column=3), int(ra_train_ct.get(ra, 0)), CENTER)
    style_cell(ws2.cell(row=ri, column=4), int(ra_est_ct.get(ra, 0)),   CENTER)

ws2.column_dimensions["A"].width = 36
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 16

wb.save(OUT)
print(f"\n✅  Saved → {OUT}")
print(f"   Sheet 'Flights':      {len(df)} rows")
print(f"   Sheet 'Model Summary': {len(ra_mae)} route×airline groups")
